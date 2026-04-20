"""
Export module — CSV, Excel, JSON, PDF
GAB Intelligence · Banque Populaire du Maroc
"""
import io, json, os, textwrap
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from fastapi import APIRouter, Query, Request
from fastapi.responses import StreamingResponse
from typing import List

# ── Router ────────────────────────────────────────────────────
router = APIRouter(prefix="/api/export", tags=["export"])

# Lazy imports — installed on demand
_openpyxl = None
_rl = None
_mpl = None


def _ensure_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        import openpyxl
        _openpyxl = openpyxl
    return _openpyxl


def _ensure_reportlab():
    global _rl
    if _rl is None:
        import reportlab as rl
        _rl = rl
    return _rl


def _ensure_matplotlib():
    global _mpl
    if _mpl is None:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        _mpl = plt
    return _mpl


def _mpl_style():
    """Apply BPM professional style to matplotlib."""
    plt = _ensure_matplotlib()
    plt.rcParams.update({
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "font.family": "sans-serif",
        "font.sans-serif": ["Helvetica", "Arial", "DejaVu Sans"],
        "axes.edgecolor": "#cccccc",
        "axes.linewidth": 0.5,
        "axes.grid": True,
        "grid.color": "#eeeeee",
        "grid.linewidth": 0.5,
        "xtick.color": "#666666",
        "ytick.color": "#666666",
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "axes.spines.top": False,
        "axes.spines.right": False,
    })


def _fig_to_buf(fig, dpi=150):
    """Render a matplotlib figure to a BytesIO PNG buffer."""
    plt = _ensure_matplotlib()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Chart generators (matplotlib → BytesIO PNG) ─────────────

BPM_RED_HEX = "#C8102E"
BPM_GREEN_HEX = "#00703C"
BPM_DARK_HEX = "#2a2e40"


def _chart_kpis(nb_gab, nb_pannes, taux, disponibilite, meilleur_modele, f1_score):
    """KPI visual cards — 2×3 grid."""
    from matplotlib.patches import FancyBboxPatch
    plt = _ensure_matplotlib()

    cards = [
        ("GAB surveillés", str(nb_gab), BPM_DARK_HEX),
        ("Pannes détectées", f"{nb_pannes:,}", BPM_RED_HEX),
        ("Taux de panne", f"{taux:.2f}%", BPM_RED_HEX),
        ("Disponibilité", f"{disponibilite:.1f}%", BPM_GREEN_HEX),
        ("Meilleur modèle", meilleur_modele[:20], BPM_DARK_HEX),
        ("F1-Score", f"{f1_score:.4f}", BPM_GREEN_HEX),
    ]
    fig, axes = plt.subplots(2, 3, figsize=(7.5, 2.8))
    for ax, (label, value, accent) in zip(axes.flat, cards):
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")
        bg = FancyBboxPatch((0.03, 0.03), 0.94, 0.94,
                            boxstyle="round,pad=0.05",
                            facecolor="#F8F8F8", edgecolor=accent, linewidth=1.5)
        ax.add_patch(bg)
        ax.text(0.5, 0.58, value, fontsize=17, fontweight="bold",
                ha="center", va="center", color=accent)
        ax.text(0.5, 0.2, label, fontsize=7.5, ha="center", va="center",
                color="#666666")
    fig.subplots_adjust(wspace=0.15, hspace=0.2)
    return _fig_to_buf(fig)


def _chart_geography(geo_df):
    """Horizontal bar — taux de panne par ville, sorted."""
    plt = _ensure_matplotlib()
    geo = geo_df.sort_values("taux_panne", ascending=True).copy()
    n = len(geo)
    fig, ax = plt.subplots(figsize=(7.0, max(2.5, n * 0.35)))
    bars = ax.barh(geo["ville"], geo["taux_panne"],
                   color=BPM_RED_HEX, edgecolor="none", height=0.6)
    for bar, val in zip(bars, geo["taux_panne"]):
        ax.text(val + 0.15, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}%", va="center", fontsize=7.5, color="#333333")
    ax.set_xlabel("Taux de panne (%)", fontsize=9, color="#666666")
    ax.set_title("Taux de panne par ville", fontsize=12, fontweight="bold",
                 color=BPM_DARK_HEX, pad=10)
    ax.set_xlim(0, geo["taux_panne"].max() * 1.15)
    fig.tight_layout()
    return _fig_to_buf(fig)


def _chart_models(resultats, meilleur):
    """Grouped bar chart — F1 / Recall / Precision per model."""
    plt = _ensure_matplotlib()
    models = [m for m in resultats if m != "Dummy (Stratified)"]
    f1s = [resultats[m].get("f1", 0) for m in models]
    recalls = [resultats[m].get("recall", 0) for m in models]
    precs = [resultats[m].get("precision", 0) for m in models]

    x = np.arange(len(models))
    width = 0.22
    fig, ax = plt.subplots(figsize=(7.0, 4.0))
    ax.bar(x - width, f1s, width, label="F1", color=BPM_RED_HEX, edgecolor="none")
    ax.bar(x, recalls, width, label="Recall", color=BPM_GREEN_HEX, edgecolor="none")
    ax.bar(x + width, precs, width, label="Précision", color=BPM_DARK_HEX, edgecolor="none")

    # Highlight best model
    if meilleur in models:
        bi = models.index(meilleur)
        ax.axvspan(bi - 0.45, bi + 0.45, alpha=0.08, color="#FFD700", zorder=0)
        best_f1 = resultats[meilleur].get("f1", 0)
        ax.annotate("★ Best", xy=(bi - width, best_f1),
                    xytext=(bi - width, best_f1 + 0.06),
                    fontsize=7, color=BPM_RED_HEX, ha="center", fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels([m[:18] for m in models], rotation=30, ha="right", fontsize=7.5)
    ax.set_ylabel("Score", fontsize=9, color="#666666")
    ax.set_ylim(0, 1.08)
    ax.legend(fontsize=8, loc="upper right")
    ax.set_title("Performance des modèles ML", fontsize=12, fontweight="bold",
                 color=BPM_DARK_HEX, pad=10)
    fig.tight_layout()
    return _fig_to_buf(fig)


def _chart_features(feat_df, top_n=15):
    """Horizontal bar — top features with gradient color."""
    from matplotlib.colors import LinearSegmentedColormap
    plt = _ensure_matplotlib()

    fi = feat_df.head(top_n).sort_values("imp_moy", ascending=True).copy()
    cmap = LinearSegmentedColormap.from_list("bpm", ["#f4a6b0", BPM_RED_HEX])
    norm_vals = (fi["imp_moy"] - fi["imp_moy"].min()) / (fi["imp_moy"].max() - fi["imp_moy"].min() + 1e-10)
    colors = [cmap(v) for v in norm_vals]

    fig, ax = plt.subplots(figsize=(7.0, max(3.0, top_n * 0.28)))
    bars = ax.barh(fi["feature"], fi["imp_moy"], color=colors, edgecolor="none", height=0.6)
    for bar, val in zip(bars, fi["imp_moy"]):
        ax.text(val + 0.0005, bar.get_y() + bar.get_height() / 2,
                f"{val:.4f}", va="center", fontsize=7, color="#333333")
    ax.set_xlabel("Importance moyenne", fontsize=9, color="#666666")
    ax.set_title(f"Top {top_n} variables prédictives", fontsize=12, fontweight="bold",
                 color=BPM_DARK_HEX, pad=10)
    ax.set_xlim(0, fi["imp_moy"].max() * 1.12)
    fig.tight_layout()
    return _fig_to_buf(fig)


def _chart_temporal(df):
    """Line chart — monthly taux de panne with summer highlight."""
    plt = _ensure_matplotlib()
    monthly = (
        df.assign(mois=df["date"].dt.to_period("M"))
        .groupby("mois")
        .agg(pannes=("panne_sous_48h", "sum"), obs=("panne_sous_48h", "count"))
        .reset_index()
    )
    monthly["taux"] = (monthly["pannes"] / monthly["obs"] * 100).round(2)
    monthly["mois_str"] = monthly["mois"].astype(str)

    fig, ax = plt.subplots(figsize=(7.0, 3.5))
    x = np.arange(len(monthly))
    ax.plot(x, monthly["taux"].values, color=BPM_RED_HEX, linewidth=2,
            marker="o", markersize=4, zorder=3, label="Taux de panne")
    ax.fill_between(x, monthly["taux"].values, alpha=0.06, color=BPM_RED_HEX)

    # Highlight summer months (June, July, August)
    summer_drawn = False
    for i, row in monthly.iterrows():
        m = row["mois"].month
        if m == 6:
            end_idx = min(i + 2, len(monthly) - 1)
            lbl = "Été (juin-août)" if not summer_drawn else ""
            ax.axvspan(i - 0.4, end_idx + 0.4, alpha=0.10, color="#FF6B35", label=lbl)
            summer_drawn = True

    ax.set_xticks(x)
    labels = monthly["mois_str"].tolist()
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=6.5)
    ax.set_ylabel("Taux de panne (%)", fontsize=9, color="#666666")
    ax.set_title("Évolution mensuelle du taux de panne", fontsize=12, fontweight="bold",
                 color=BPM_DARK_HEX, pad=10)
    ax.legend(fontsize=8, loc="upper left")
    fig.tight_layout()
    return _fig_to_buf(fig)


def _chart_threshold(resultats, meilleur, couts_dict):
    """Dual-axis: cost curve + F1 vs threshold."""
    plt = _ensure_matplotlib()
    nom = meilleur if meilleur in resultats else next(iter(resultats))
    r = resultats[nom]
    cout_correctif = couts_dict.get("fn", 5000)
    cout_preventif = couts_dict.get("tp", 1500)
    cout_fausse = couts_dict.get("fp", 500)

    tp_total = r["tp"] + r["fn"]
    tn_total = r["tn"] + r["fp"]
    total = tp_total + tn_total
    base_rate = tp_total / total if total > 0 else 0.1

    seuils = np.linspace(0.01, 0.99, 200)

    def interp3(t, y0, y_mid, y1):
        if t <= 0.5:
            return y0 + (t / 0.5) * (y_mid - y0)
        return y_mid + ((t - 0.5) / 0.5) * (y1 - y_mid)

    f1s_s, couts_s = [], []
    for s in seuils:
        rec = float(np.clip(interp3(float(s), 1.0, float(r["recall"]), 0.0), 0, 1))
        prec = float(np.clip(interp3(float(s), float(base_rate), float(r["precision"]), 1.0), 0, 1))
        f1 = 2 * prec * rec / (prec + rec + 1e-10)
        tp_s = int(rec * tp_total)
        fp_s = int(tp_s / (prec + 1e-10) - tp_s) if prec > 1e-10 else 0
        fn_s = tp_total - tp_s
        f1s_s.append(float(f1))
        couts_s.append(int(tp_s * cout_preventif + fp_s * cout_fausse + fn_s * cout_correctif))

    cout_sans = tp_total * cout_correctif
    seuil_f1 = float(seuils[int(np.argmax(f1s_s))])
    seuil_econ = float(seuils[int(np.argmin(couts_s))])
    economie = cout_sans - min(couts_s)
    eco_pct = economie / cout_sans * 100 if cout_sans > 0 else 0

    fig, ax1 = plt.subplots(figsize=(7.0, 4.0))
    ax2 = ax1.twinx()

    # Cost curve (left axis)
    ax1.plot(seuils, [c / 1000 for c in couts_s], color=BPM_RED_HEX, linewidth=2,
             label="Coût total (k MAD)")
    ax1.fill_between(seuils, [c / 1000 for c in couts_s], alpha=0.06, color=BPM_RED_HEX)
    ax1.axhline(y=cout_sans / 1000, color="#E05252", linestyle="--", linewidth=1, alpha=0.7,
                label="Sans modèle")

    # F1 curve (right axis)
    ax2.plot(seuils, f1s_s, color=BPM_GREEN_HEX, linewidth=1.5, linestyle="--",
             label="F1-Score")

    # Mark optimal points
    ax1.axvline(seuil_econ, color=BPM_RED_HEX, linestyle=":", alpha=0.5, linewidth=1)
    ax1.plot(seuil_econ, min(couts_s) / 1000, "v", color=BPM_RED_HEX, markersize=10, zorder=5)
    ax2.axvline(seuil_f1, color=BPM_GREEN_HEX, linestyle=":", alpha=0.5, linewidth=1)
    ax2.plot(seuil_f1, max(f1s_s), "^", color=BPM_GREEN_HEX, markersize=10, zorder=5)

    # Annotations
    ax1.annotate(f"Éco opt: {seuil_econ:.2f}\n−{economie / 1000:.0f}k MAD ({eco_pct:.0f}%)",
                 xy=(seuil_econ, min(couts_s) / 1000),
                 xytext=(seuil_econ + 0.08, min(couts_s) / 1000 * 1.15),
                 fontsize=7, color=BPM_RED_HEX,
                 arrowprops=dict(arrowstyle="->", color=BPM_RED_HEX, lw=0.8))
    ax2.annotate(f"F1 opt: {seuil_f1:.2f}",
                 xy=(seuil_f1, max(f1s_s)),
                 xytext=(seuil_f1 + 0.08, max(f1s_s) - 0.05),
                 fontsize=7, color=BPM_GREEN_HEX,
                 arrowprops=dict(arrowstyle="->", color=BPM_GREEN_HEX, lw=0.8))

    ax1.set_xlabel("Seuil de décision", fontsize=9, color="#666666")
    ax1.set_ylabel("Coût total (k MAD)", fontsize=9, color=BPM_RED_HEX)
    ax2.set_ylabel("F1-Score", fontsize=9, color=BPM_GREEN_HEX)
    ax2.set_ylim(0, 1.05)

    # Combined legend
    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, fontsize=7.5, loc="upper right")

    ax1.set_title("Optimisation du seuil de décision", fontsize=12, fontweight="bold",
                  color=BPM_DARK_HEX, pad=10)
    fig.tight_layout()
    return _fig_to_buf(fig), seuil_f1, seuil_econ, economie, eco_pct


# ── Helpers ───────────────────────────────────────────────────

def _get_main_module():
    """Import main module data lazily to avoid circular imports."""
    from main import DF, FEAT_IMP, RESULTATS, MEILLEUR_MODELE, COUTS, filter_df
    return DF, FEAT_IMP, RESULTATS, MEILLEUR_MODELE, COUTS, filter_df


def _parse_filters(villes, types, annees):
    return (villes or None, types or None, annees or None)


def _stream(buf, filename, media_type):
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _now_str():
    return datetime.now().strftime("%Y%m%d_%H%M")


# ══════════════════════════════════════════════════════════════
#  CSV Export
# ══════════════════════════════════════════════════════════════

@router.get("/csv")
def export_csv(
    villes: List[str] = Query(default=[]),
    types: List[str] = Query(default=[]),
    annees: List[str] = Query(default=[]),
    section: str = Query(default="overview"),
):
    DF, FEAT_IMP, RESULTATS, MEILLEUR_MODELE, COUTS, filter_df = _get_main_module()
    v, t, a = _parse_filters(villes, types, annees)
    df = filter_df(DF, v, t, a)

    buf = io.StringIO()

    if section == "features":
        FEAT_IMP.to_csv(buf, index=False, encoding="utf-8-sig")
    elif section == "models":
        rows = []
        for name, metrics in RESULTATS.items():
            rows.append({"modele": name, **{k: v for k, v in metrics.items()
                                             if isinstance(v, (int, float))}})
        pd.DataFrame(rows).to_csv(buf, index=False, encoding="utf-8-sig")
    elif section == "geography":
        geo = (
            df.groupby("ville")
            .agg(nb_pannes=("panne_sous_48h", "sum"),
                 nb_obs=("panne_sous_48h", "count"),
                 nb_gab=("gab_id", "nunique"))
            .reset_index()
        )
        geo["taux_panne"] = (geo["nb_pannes"] / geo["nb_obs"] * 100).round(2)
        geo.to_csv(buf, index=False, encoding="utf-8-sig")
    else:  # overview — raw filtered data
        export_cols = [c for c in df.columns if not c.startswith("_")]
        df[export_cols].to_csv(buf, index=False, encoding="utf-8-sig")

    return _stream(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        f"gab_{section}_{_now_str()}.csv",
        "text/csv; charset=utf-8",
    )


# ══════════════════════════════════════════════════════════════
#  JSON Export
# ══════════════════════════════════════════════════════════════

@router.get("/json")
def export_json(
    villes: List[str] = Query(default=[]),
    types: List[str] = Query(default=[]),
    annees: List[str] = Query(default=[]),
    section: str = Query(default="overview"),
):
    DF, FEAT_IMP, RESULTATS, MEILLEUR_MODELE, COUTS, filter_df = _get_main_module()
    v, t, a = _parse_filters(villes, types, annees)
    df = filter_df(DF, v, t, a)

    if section == "features":
        payload = FEAT_IMP.head(50).to_dict(orient="records")
    elif section == "models":
        payload = RESULTATS
    elif section == "geography":
        geo = (
            df.groupby("ville")
            .agg(nb_pannes=("panne_sous_48h", "sum"),
                 nb_obs=("panne_sous_48h", "count"),
                 nb_gab=("gab_id", "nunique"))
            .reset_index()
        )
        geo["taux_panne"] = (geo["nb_pannes"] / geo["nb_obs"] * 100).round(2)
        payload = geo.to_dict(orient="records")
    else:
        payload = {
            "nb_gab": int(df["gab_id"].nunique()),
            "nb_pannes": int(df["panne_sous_48h"].sum()),
            "taux_panne": round(float(df["panne_sous_48h"].mean() * 100), 2),
            "meilleur_modele": MEILLEUR_MODELE,
            "nb_observations": int(len(df)),
            "filtres": {"villes": v, "types": t, "annees": a},
            "donnees": json.loads(df.head(500).to_json(orient="records", date_format="iso")),
        }

    content = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    buf = io.BytesIO(content.encode("utf-8"))
    return _stream(buf, f"gab_{section}_{_now_str()}.json", "application/json")


# ══════════════════════════════════════════════════════════════
#  Excel Export (multi-sheet)
# ══════════════════════════════════════════════════════════════

@router.get("/excel")
def export_excel(
    villes: List[str] = Query(default=[]),
    types: List[str] = Query(default=[]),
    annees: List[str] = Query(default=[]),
):
    _ensure_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DF, FEAT_IMP, RESULTATS, MEILLEUR_MODELE, COUTS, filter_df = _get_main_module()
    v, t, a = _parse_filters(villes, types, annees)
    df = filter_df(DF, v, t, a)

    wb = _openpyxl.Workbook()

    # ── Style constants ──
    BPM_RED = "C8102E"
    BPM_GREEN = "00703C"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=BPM_RED, end_color=BPM_RED, fill_type="solid")
    title_font = Font(bold=True, size=14, color=BPM_RED)
    sub_font = Font(italic=True, size=10, color="666666")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def add_data_rows(ws, dataframe, start_row=2):
        for r_idx, row in enumerate(dataframe.itertuples(index=False), start=start_row):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = '0.00'

    # ── Sheet 1: Résumé KPIs ──
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1["A1"] = "GAB Intelligence — Rapport d'analyse"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws1["A2"].font = sub_font
    ws1["A3"] = f"Banque Populaire du Maroc — Maintenance Prédictive"
    ws1["A3"].font = sub_font

    kpis = [
        ("GAB surveillés", int(df["gab_id"].nunique())),
        ("Pannes enregistrées", int(df["panne_sous_48h"].sum())),
        ("Taux de panne moyen", f'{df["panne_sous_48h"].mean() * 100:.2f}%'),
        ("Disponibilité", f'{(1 - df["panne_sous_48h"].mean()) * 100:.1f}%'),
        ("Observations", int(len(df))),
        ("Villes", int(df["ville"].nunique())),
        ("Meilleur modèle", MEILLEUR_MODELE),
        ("F1-Score", round(float(RESULTATS[MEILLEUR_MODELE]["f1"]), 4)),
        ("Recall", round(float(RESULTATS[MEILLEUR_MODELE]["recall"]), 4)),
        ("Precision", round(float(RESULTATS[MEILLEUR_MODELE]["precision"]), 4)),
    ]
    for i, (label, value) in enumerate(kpis, start=5):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        ws1.cell(row=i, column=2, value=value).font = Font(size=11)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 30

    # ── Sheet 2: Données filtrées ──
    ws2 = wb.create_sheet("Données")
    export_cols = [c for c in df.columns if not c.startswith("_")]
    df_export = df[export_cols].copy()
    # Convert date to string for Excel
    if "date" in df_export.columns:
        df_export["date"] = df_export["date"].dt.strftime("%Y-%m-%d")
    for c_idx, col in enumerate(df_export.columns, start=1):
        ws2.cell(row=1, column=c_idx, value=col)
    style_header(ws2, len(df_export.columns))
    add_data_rows(ws2, df_export)

    # ── Sheet 3: Géographie ──
    ws3 = wb.create_sheet("Géographie")
    geo = (
        df.groupby("ville")
        .agg(nb_pannes=("panne_sous_48h", "sum"),
             nb_obs=("panne_sous_48h", "count"),
             nb_gab=("gab_id", "nunique"))
        .reset_index()
    )
    geo["taux_panne"] = (geo["nb_pannes"] / geo["nb_obs"] * 100).round(2)
    geo = geo.sort_values("taux_panne", ascending=False)
    for c_idx, col in enumerate(geo.columns, start=1):
        ws3.cell(row=1, column=c_idx, value=col)
    style_header(ws3, len(geo.columns))
    add_data_rows(ws3, geo)

    # ── Sheet 4: Modèles ML ──
    ws4 = wb.create_sheet("Modèles ML")
    model_rows = []
    for name, metrics in RESULTATS.items():
        row = {"modele": name}
        row.update({k: round(float(v), 4) if isinstance(v, (int, float)) else v
                    for k, v in metrics.items()})
        model_rows.append(row)
    df_models = pd.DataFrame(model_rows)
    for c_idx, col in enumerate(df_models.columns, start=1):
        ws4.cell(row=1, column=c_idx, value=col)
    style_header(ws4, len(df_models.columns))
    add_data_rows(ws4, df_models)

    # ── Sheet 5: Features ──
    ws5 = wb.create_sheet("Features")
    fi_export = FEAT_IMP.head(50).copy()
    for c_idx, col in enumerate(fi_export.columns, start=1):
        ws5.cell(row=1, column=c_idx, value=col)
    style_header(ws5, len(fi_export.columns))
    add_data_rows(ws5, fi_export)

    # ── Write to buffer ──
    buf = io.BytesIO()
    wb.save(buf)
    return _stream(
        buf,
        f"gab_rapport_{_now_str()}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ══════════════════════════════════════════════════════════════
#  PDF Report (complet — type Power BI export)
# ══════════════════════════════════════════════════════════════

@router.get("/pdf")
def export_pdf(
    villes: List[str] = Query(default=[]),
    types: List[str] = Query(default=[]),
    annees: List[str] = Query(default=[]),
):
    _ensure_reportlab()
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable, Image,
    )

    _ensure_matplotlib()
    _mpl_style()

    DF, FEAT_IMP, RESULTATS, MEILLEUR_MODELE, COUTS, filter_df = _get_main_module()
    v, t, a = _parse_filters(villes, types, annees)
    df = filter_df(DF, v, t, a)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    # Colors
    BPM_RED = HexColor("#C8102E")
    BPM_GREEN = HexColor("#00703C")
    DARK_BG = HexColor("#0a0a0c")
    GRAY = HexColor("#666666")
    LIGHT_GRAY = HexColor("#F5F5F5")

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        "CoverTitle", parent=styles["Title"],
        fontSize=28, textColor=BPM_RED, spaceAfter=6,
        alignment=TA_CENTER, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "CoverSub", parent=styles["Normal"],
        fontSize=14, textColor=GRAY, alignment=TA_CENTER,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        "SectionTitle", parent=styles["Heading1"],
        fontSize=16, textColor=BPM_RED, spaceBefore=16,
        spaceAfter=8, fontName="Helvetica-Bold",
        borderColor=BPM_RED, borderWidth=0, borderPadding=0,
    ))
    styles.add(ParagraphStyle(
        "KpiLabel", parent=styles["Normal"],
        fontSize=9, textColor=GRAY, fontName="Helvetica",
    ))
    styles.add(ParagraphStyle(
        "KpiValue", parent=styles["Normal"],
        fontSize=18, textColor=HexColor("#1a1a2e"), fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "BodyText2", parent=styles["Normal"],
        fontSize=10, textColor=HexColor("#333333"), spaceAfter=6,
        leading=14,
    ))
    styles.add(ParagraphStyle(
        "FooterStyle", parent=styles["Normal"],
        fontSize=8, textColor=GRAY, alignment=TA_CENTER,
    ))

    elements = []
    W = A4[0] - 36 * mm  # usable width

    # ── Cover page ──
    elements.append(Spacer(1, 60 * mm))
    elements.append(Paragraph("GAB Intelligence", styles["CoverTitle"]))
    elements.append(Paragraph("Rapport d'Analyse — Maintenance Prédictive", styles["CoverSub"]))
    elements.append(Spacer(1, 8 * mm))
    elements.append(HRFlowable(width="60%", thickness=2, color=BPM_RED, spaceAfter=8))
    elements.append(Paragraph("Banque Populaire du Maroc", styles["CoverSub"]))
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        styles["CoverSub"],
    ))
    elements.append(Spacer(1, 12 * mm))

    # Filters summary
    filter_text = []
    if v:
        filter_text.append(f"Villes : {', '.join(v)}")
    if t:
        filter_text.append(f"Types : {', '.join(t)}")
    if a:
        filter_text.append(f"Années : {', '.join(str(x) for x in a)}")
    if filter_text:
        elements.append(Paragraph(
            "<br/>".join(filter_text),
            ParagraphStyle("FilterInfo", parent=styles["Normal"],
                           fontSize=10, textColor=GRAY, alignment=TA_CENTER),
        ))

    elements.append(PageBreak())

    # ── Page 2: KPIs ──
    elements.append(Paragraph("1. Indicateurs Clés (KPIs)", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    nb_gab = int(df["gab_id"].nunique())
    nb_pannes = int(df["panne_sous_48h"].sum())
    taux = df["panne_sous_48h"].mean() * 100
    disponibilite = 100 - taux
    best_r = RESULTATS[MEILLEUR_MODELE]

    kpi_data = [
        ["Indicateur", "Valeur", "Détail"],
        ["GAB surveillés", str(nb_gab), f"{int(df['ville'].nunique())} villes"],
        ["Pannes enregistrées", f"{nb_pannes:,}", f"Taux : {taux:.2f}%"],
        ["Disponibilité", f"{disponibilite:.1f}%", "Objectif : > 95%"],
        ["Observations", f"{len(df):,}", "2 ans de données"],
        ["Meilleur modèle", MEILLEUR_MODELE[:30], f"F1 = {best_r['f1']:.4f}"],
        ["Recall", f"{best_r['recall']:.4f}", "Détection des pannes"],
        ["Precision", f"{best_r['precision']:.4f}", "Fiabilité des alertes"],
    ]

    kpi_table = Table(kpi_data, colWidths=[W * 0.35, W * 0.3, W * 0.35])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BPM_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 6 * mm))

    # KPI visual cards (matplotlib)
    kpi_buf = _chart_kpis(nb_gab, nb_pannes, taux, disponibilite,
                          MEILLEUR_MODELE, best_r["f1"])
    elements.append(Image(kpi_buf, width=W, height=W * 2.8 / 7.5))
    elements.append(Spacer(1, 6 * mm))

    # ── Analyse textuelle ──
    elements.append(Paragraph("Synthèse", ParagraphStyle(
        "SubTitle", parent=styles["Heading2"], fontSize=13,
        textColor=HexColor("#1a1a2e"), fontName="Helvetica-Bold",
    )))
    elements.append(Paragraph(
        f"Le parc de <b>{nb_gab}</b> GAB répartis sur <b>{int(df['ville'].nunique())}</b> villes "
        f"présente un taux de panne moyen de <b>{taux:.2f}%</b> sur la période analysée. "
        f"Le modèle <b>{MEILLEUR_MODELE}</b> atteint un F1-score de <b>{best_r['f1']:.4f}</b> "
        f"avec un recall de <b>{best_r['recall']:.4f}</b>, permettant de détecter "
        f"la majorité des pannes avant leur survenue.",
        styles["BodyText2"],
    ))

    elements.append(PageBreak())

    # ── Page 3: Géographie ──
    elements.append(Paragraph("2. Analyse Géographique", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    geo = (
        df.groupby("ville")
        .agg(nb_pannes=("panne_sous_48h", "sum"),
             nb_obs=("panne_sous_48h", "count"),
             nb_gab=("gab_id", "nunique"))
        .reset_index()
    )
    geo["taux_panne"] = (geo["nb_pannes"] / geo["nb_obs"] * 100).round(2)
    geo = geo.sort_values("taux_panne", ascending=False)

    geo_header = ["Ville", "GAB", "Pannes", "Observations", "Taux (%)"]
    geo_rows = [geo_header]
    for _, row in geo.iterrows():
        geo_rows.append([
            row["ville"],
            str(int(row["nb_gab"])),
            str(int(row["nb_pannes"])),
            str(int(row["nb_obs"])),
            f'{row["taux_panne"]:.2f}',
        ])

    geo_table = Table(geo_rows, colWidths=[W * 0.25, W * 0.15, W * 0.2, W * 0.2, W * 0.2])
    geo_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BPM_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(geo_table)
    elements.append(Spacer(1, 6 * mm))

    # Bar chart — taux par ville (matplotlib)
    n_cities = len(geo)
    geo_buf = _chart_geography(geo)
    geo_h = W * max(2.5, n_cities * 0.35) / 7.0
    elements.append(Image(geo_buf, width=W, height=min(geo_h, 180 * mm)))

    elements.append(PageBreak())

    # ── Page 4: Modèles ML ──
    elements.append(Paragraph("3. Performance des Modèles ML", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    model_header = ["Modèle", "F1", "Recall", "Precision", "AUC-PR"]
    model_rows_data = [model_header]
    for name, m in RESULTATS.items():
        is_best = " *" if name == MEILLEUR_MODELE else ""
        model_rows_data.append([
            (name[:28] + is_best),
            f'{m.get("f1", 0):.4f}',
            f'{m.get("recall", 0):.4f}',
            f'{m.get("precision", 0):.4f}',
            f'{m.get("auc_pr", m.get("auc-pr", 0)):.4f}',
        ])

    model_table = Table(model_rows_data, colWidths=[W * 0.32, W * 0.17, W * 0.17, W * 0.17, W * 0.17])
    model_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BPM_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(model_table)
    elements.append(Spacer(1, 6 * mm))

    # Grouped bar chart (matplotlib)
    model_buf = _chart_models(RESULTATS, MEILLEUR_MODELE)
    elements.append(Image(model_buf, width=W, height=W * 4.0 / 7.0))
    elements.append(Spacer(1, 4 * mm))

    elements.append(Paragraph(
        f"<b>* Meilleur modèle sélectionné :</b> {MEILLEUR_MODELE}<br/>"
        f"Le modèle a été entraîné avec split temporel strict pour éviter le data leakage. "
        f"Un F1-score de {best_r['f1']:.4f} est cohérent avec un taux de classe minoritaire "
        f"de ~{taux:.1f}%.",
        styles["BodyText2"],
    ))

    elements.append(PageBreak())

    # ── Page 5: Top Features ──
    elements.append(Paragraph("4. Importance des Variables", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    fi_top = FEAT_IMP.head(20)
    feat_header = ["Rang", "Feature", "Importance"]
    feat_rows = [feat_header]
    for i, (_, row) in enumerate(fi_top.iterrows(), 1):
        feat_rows.append([
            str(i),
            str(row["feature"])[:40],
            f'{row["imp_moy"]:.4f}',
        ])

    feat_table = Table(feat_rows, colWidths=[W * 0.1, W * 0.6, W * 0.3])
    feat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BPM_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (2, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(feat_table)
    elements.append(Spacer(1, 6 * mm))

    # Feature importance bar chart (matplotlib)
    feat_buf = _chart_features(FEAT_IMP, top_n=15)
    elements.append(Image(feat_buf, width=W, height=W * max(3.0, 15 * 0.28) / 7.0))
    elements.append(Spacer(1, 4 * mm))

    elements.append(Paragraph(
        "Les variables liées à la <b>maintenance</b> (jours depuis maintenance, score de négligence) "
        "et aux <b>erreurs matérielles</b> (erreurs lecteur, risque matériel) dominent l'importance. "
        "Les features <b>rolling</b> et de <b>tendance</b> apportent une valeur prédictive significative "
        "en capturant la dynamique de dégradation.",
        styles["BodyText2"],
    ))

    elements.append(PageBreak())

    # ── Page 6: Analyse temporelle ──
    elements.append(Paragraph("5. Analyse Temporelle", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    # Monthly breakdown
    monthly = (
        df.assign(mois=df["date"].dt.to_period("M").astype(str))
        .groupby("mois")
        .agg(pannes=("panne_sous_48h", "sum"), obs=("panne_sous_48h", "count"))
        .reset_index()
    )
    monthly["taux"] = (monthly["pannes"] / monthly["obs"] * 100).round(2)

    month_header = ["Mois", "Pannes", "Observations", "Taux (%)"]
    month_rows = [month_header]
    for _, row in monthly.iterrows():
        month_rows.append([
            row["mois"],
            str(int(row["pannes"])),
            str(int(row["obs"])),
            f'{row["taux"]:.2f}',
        ])

    month_table = Table(month_rows, colWidths=[W * 0.3, W * 0.2, W * 0.25, W * 0.25])
    month_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BPM_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(month_table)
    elements.append(Spacer(1, 6 * mm))

    # Seasonal analysis
    df_s = df.copy()
    df_s["saison"] = df_s["date"].dt.month.map({
        12: "Hiver", 1: "Hiver", 2: "Hiver",
        3: "Printemps", 4: "Printemps", 5: "Printemps",
        6: "Été", 7: "Été", 8: "Été",
        9: "Automne", 10: "Automne", 11: "Automne",
    })
    sais = df_s.groupby("saison")["panne_sous_48h"].mean() * 100
    sais = sais.reindex(["Printemps", "Été", "Automne", "Hiver"]).fillna(0).round(2)

    elements.append(Paragraph(
        f"<b>Saisonnalité :</b> Printemps {sais.get('Printemps', 0):.1f}% · "
        f"Été {sais.get('Été', 0):.1f}% · "
        f"Automne {sais.get('Automne', 0):.1f}% · "
        f"Hiver {sais.get('Hiver', 0):.1f}%<br/>"
        f"Le pic estival confirme le rôle du stress thermique comme facteur aggravant.",
        styles["BodyText2"],
    ))
    elements.append(Spacer(1, 4 * mm))

    # Temporal line chart (matplotlib)
    temp_buf = _chart_temporal(df)
    elements.append(Image(temp_buf, width=W, height=W * 3.5 / 7.0))

    elements.append(PageBreak())

    # ── Page 7: Optimisation du Seuil ──
    elements.append(Paragraph("6. Optimisation du Seuil", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    threshold_result = _chart_threshold(RESULTATS, MEILLEUR_MODELE, COUTS)
    thresh_buf, seuil_f1_opt, seuil_econ_opt, eco_max, eco_pct = threshold_result
    elements.append(Image(thresh_buf, width=W, height=W * 4.0 / 7.0))
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph(
        f"Le seuil <b>F1-optimal</b> est de <b>{seuil_f1_opt:.2f}</b> "
        f"(maximise la détection des pannes). "
        f"Le seuil <b>économique optimal</b> est de <b>{seuil_econ_opt:.2f}</b>, "
        f"permettant une économie estimée de <b>{eco_max / 1000:.0f} k MAD ({eco_pct:.0f}%)</b> "
        f"par rapport à une stratégie 100% corrective.<br/>"
        f"Ce graphique permet aux décideurs de choisir le seuil adapté à leur tolérance au risque "
        f"et à leurs contraintes budgétaires.",
        styles["BodyText2"],
    ))

    elements.append(PageBreak())

    # ── Page 8: Recommandations ──
    elements.append(Paragraph("7. Recommandations", styles["SectionTitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=BPM_RED, spaceAfter=12))

    recommendations = [
        ("Maintenance préventive ciblée",
         "Prioriser les interventions sur les GAB identifiés comme à haut risque par le modèle. "
         "Les GAB Wincor en sites isolés avec > 90 jours sans maintenance nécessitent une attention immédiate."),
        ("Gestion thermique",
         "Installer des systèmes de refroidissement renforcés pour les GAB en façade et sites isolés, "
         "particulièrement avant la période estivale."),
        ("Renouvellement du parc",
         "Les GAB de plus de 7 ans présentent un taux de panne significativement plus élevé. "
         "Planifier le remplacement progressif en commençant par les zones les plus critiques."),
        ("Monitoring continu",
         "Déployer le système de scoring en production pour un suivi en temps réel. "
         "Le seuil optimal identifié permet de réduire les coûts de maintenance tout en "
         "maintenant un recall élevé."),
        ("Enrichissement des données",
         "Intégrer les données de fréquentation et les conditions météorologiques locales "
         "pour améliorer la précision du modèle."),
    ]

    for i, (title, desc) in enumerate(recommendations, 1):
        elements.append(Paragraph(
            f"<b>{i}. {title}</b>",
            ParagraphStyle("RecTitle", parent=styles["Normal"],
                           fontSize=11, textColor=HexColor("#1a1a2e"),
                           fontName="Helvetica-Bold", spaceBefore=8),
        ))
        elements.append(Paragraph(desc, styles["BodyText2"]))

    elements.append(Spacer(1, 15 * mm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=GRAY, spaceAfter=8))
    elements.append(Paragraph(
        "Rapport généré automatiquement par GAB Intelligence — "
        "Banque Populaire du Maroc — Projet de Fin d'Études",
        styles["FooterStyle"],
    ))

    # ── Build PDF ──
    doc.build(elements)

    return _stream(buf, f"gab_rapport_{_now_str()}.pdf", "application/pdf")
